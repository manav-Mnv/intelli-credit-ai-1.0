"""
Generate a sample Excel file with financial data for demo.
Usage: python sample_data\generate_sample_excel.py
"""
import os

def generate():
    os.makedirs("sample_data", exist_ok=True)

    try:
        import openpyxl
        wb = openpyxl.Workbook()

        # --- P&L Sheet ---
        ws = wb.active
        ws.title = "Profit & Loss"
        headers = ["Particulars", "FY 2024 (₹)", "FY 2023 (₹)"]
        ws.append(headers)
        data = [
            ["Revenue from Operations", 650000000, 580000000],
            ["Cost of Materials", 350000000, 310000000],
            ["Employee Benefits", 45000000, 42000000],
            ["Finance Cost / Interest Expense", 21000000, 19000000],
            ["Depreciation", 28000000, 25000000],
            ["Other Expenses", 85000000, 80000000],
            ["Total Expenses", 529000000, 476000000],
            ["Profit Before Tax", 126000000, 108500000],
            ["Tax Expense", 56000000, 48500000],
            ["Net Profit", 70000000, 60000000],
            ["EBITDA", 105000000, 93000000],
        ]
        for row in data:
            ws.append(row)

        # --- Balance Sheet ---
        bs = wb.create_sheet("Balance Sheet")
        bs.append(["Particulars", "FY 2024 (₹)", "FY 2023 (₹)"])
        bs_data = [
            ["Share Capital", 50000000, 50000000],
            ["Reserves and Surplus", 200000000, 180000000],
            ["Total Equity", 250000000, 230000000],
            ["Long Term Borrowings", 120000000, 100000000],
            ["Short Term Borrowings", 60000000, 55000000],
            ["Total Debt", 180000000, 155000000],
            ["Trade Payables", 55000000, 48000000],
            ["Other Current Liabilities", 25000000, 22000000],
            ["Total Current Liabilities", 140000000, 125000000],
            ["Property Plant Equipment", 180000000, 160000000],
            ["Total Assets", 450000000, 420000000],
            ["Current Assets", 220000000, 196000000],
            ["Cash and Bank Balances", 25000000, 20000000],
        ]
        for row in bs_data:
            bs.append(row)

        # --- Ratios Sheet ---
        rs = wb.create_sheet("Key Ratios")
        rs.append(["Ratio", "FY 2024", "FY 2023"])
        ratios = [
            ["Debt to Equity Ratio", 0.72, 0.65],
            ["Current Ratio", 1.57, 1.44],
            ["Profit Margin (%)", 10.77, 10.26],
            ["Return on Equity (%)", 28.0, 26.09],
            ["Interest Coverage Ratio", 5.0, 4.68],
            ["EBITDA Margin (%)", 16.15, 16.03],
        ]
        for row in ratios:
            rs.append(row)

        out_path = os.path.join("sample_data", "ABC_Steel_Financials_2024.xlsx")
        wb.save(out_path)
        print(f"✅ Sample Excel created: {out_path}")
        return out_path

    except ImportError:
        print("❌ openpyxl not installed. Run: pip install openpyxl")
        return None


if __name__ == "__main__":
    generate()
