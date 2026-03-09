"""
Document Processing Engine
Handles PDF, Excel, Word documents using PyMuPDF, pdfplumber, and openpyxl.
Extracts text and financial tables, returning structured financial data.
"""

import logging
import re
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Main document processing engine with multiple extraction strategies."""

    def __init__(self):
        self._pymupdf_available = self._check_pymupdf()
        self._pdfplumber_available = self._check_pdfplumber()
        self._openpyxl_available = self._check_openpyxl()

    def _check_pymupdf(self) -> bool:
        try:
            import fitz  # noqa
            return True
        except ImportError:
            logger.warning("PyMuPDF not available")
            return False

    def _check_pdfplumber(self) -> bool:
        try:
            import pdfplumber  # noqa
            return True
        except ImportError:
            logger.warning("pdfplumber not available")
            return False

    def _check_openpyxl(self) -> bool:
        try:
            import openpyxl  # noqa
            return True
        except ImportError:
            logger.warning("openpyxl not available")
            return False

    def process_file(self, file_path: str) -> dict:
        """Process any supported document type and extract financial data."""
        path = Path(file_path)
        ext = path.suffix.lower()

        logger.info(f"Processing document: {path.name} (type: {ext})")
        text = ""

        if ext == ".pdf":
            text = self._extract_pdf_text(file_path)
        elif ext in (".xlsx", ".xls"):
            text = self._extract_excel_data(file_path)
        elif ext in (".docx", ".doc"):
            text = self._extract_word_text(file_path)
        elif ext == ".txt":
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
        else:
            logger.warning(f"Unsupported file type: {ext}")

        financial_data = self._extract_financial_data(text)
        return {
            "raw_text": text[:5000] if text else "",
            "page_count": self._count_pages(file_path, ext),
            "financial_data": financial_data,
            "extraction_method": self._get_extraction_method(ext),
        }

    def _extract_pdf_text(self, file_path: str) -> str:
        """Extract text from PDF using PyMuPDF with pdfplumber fallback."""
        text = ""

        if self._pymupdf_available:
            try:
                import fitz
                doc = fitz.open(file_path)
                for page in doc:
                    text += page.get_text("text") + "\n"
                doc.close()
                logger.info(f"PyMuPDF extracted {len(text)} chars")
                return text
            except Exception as e:
                logger.warning(f"PyMuPDF failed: {e}, trying pdfplumber")

        if self._pdfplumber_available:
            try:
                import pdfplumber
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                        # Extract tables
                        tables = page.extract_tables()
                        for table in tables:
                            for row in table:
                                text += " | ".join(str(cell) for cell in row if cell) + "\n"
                logger.info(f"pdfplumber extracted {len(text)} chars")
                return text
            except Exception as e:
                logger.error(f"pdfplumber failed: {e}")

        return text

    def _extract_excel_data(self, file_path: str) -> str:
        """Extract data from Excel files."""
        if not self._openpyxl_available:
            return ""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"\n=== Sheet: {sheet} ===\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        text += row_text + "\n"
            return text
        except Exception as e:
            logger.error(f"Excel extraction failed: {e}")
            return ""

    def _extract_word_text(self, file_path: str) -> str:
        """Extract text from Word documents."""
        try:
            from docx import Document
            doc = Document(file_path)
            return "\n".join(para.text for para in doc.paragraphs)
        except Exception as e:
            logger.error(f"Word extraction failed: {e}")
            return ""

    def _count_pages(self, file_path: str, ext: str) -> int:
        if ext == ".pdf" and self._pymupdf_available:
            try:
                import fitz
                doc = fitz.open(file_path)
                count = len(doc)
                doc.close()
                return count
            except Exception:
                pass
        return 1

    def _get_extraction_method(self, ext: str) -> str:
        if ext == ".pdf":
            if self._pymupdf_available:
                return "pymupdf"
            elif self._pdfplumber_available:
                return "pdfplumber"
        elif ext in (".xlsx", ".xls"):
            return "openpyxl"
        return "text"

    def _extract_financial_data(self, text: str) -> dict:
        """
        Extract key financial figures from text using regex pattern matching.
        Handles: revenue, profit, debt, equity, assets, liabilities.
        Numbers can be in crores (₹Cr), lakhs, or raw.
        """
        text_lower = text.lower()
        data = {}

        patterns = {
            "revenue": [
                r"(?:total\s+)?revenue[s]?\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
                r"(?:net\s+)?sales\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
                r"turnover\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
            ],
            "net_profit": [
                r"(?:net\s+)?profit(?:\s+after\s+tax)?\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
                r"pat\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
                r"net\s+income\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
            ],
            "total_debt": [
                r"(?:total\s+)?(?:long[\s\-]term\s+)?debt\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
                r"borrowings\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
                r"total\s+liabilities\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
            ],
            "total_equity": [
                r"(?:shareholders?[''s]?\s+)?equity\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
                r"net\s+worth\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
            ],
            "total_assets": [
                r"total\s+assets\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
                r"balance\s+sheet\s+total\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
            ],
            "ebitda": [
                r"ebitda\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
            ],
            "interest_expense": [
                r"interest\s+(?:expense|cost|paid)\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
                r"finance\s+cost[s]?\s*[:\-]?\s*₹?\s*([\d,\.]+)\s*(?:cr|crore|lakh)?",
            ],
        }

        for field, field_patterns in patterns.items():
            for pattern in field_patterns:
                match = re.search(pattern, text_lower)
                if match:
                    try:
                        raw = match.group(1).replace(",", "")
                        value = float(raw)
                        # Convert crores to actual value (×10M)
                        if "cr" in text_lower[max(0, match.start()-5):match.end()+10]:
                            value = value * 10_000_000
                        elif "lakh" in text_lower[max(0, match.start()-5):match.end()+10]:
                            value = value * 100_000
                        data[field] = value
                    except ValueError:
                        pass
                    break  # Stop after first match for this field

        return data


# Singleton instance
processor = DocumentProcessor()
